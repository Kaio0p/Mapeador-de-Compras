# -*- coding: utf-8 -*-
"""
excel_generator.py
Preenche o template real do Mapa de Compras.
Carrega template/mapa_compras_template.xlsx e escreve apenas nas células de dados.
Toda formatação, fórmulas e estrutura vêm do template original intocados.

Mapeamento do template:
  Cabeçalho:
    C6 = Número Sequencial | F6 = Filial | H6 = Responsável | R6 = Data

  Fornecedores (linha 10): I10 J10 K10 L10

  Dados (linhas 11-53):
    E=Item  F=Marca  G=Qtd  H=UND
    I=Forn1  J=Forn2  K=Forn3  L=Forn4
    P=Preço Autorizado (preenchido quando um fornecedor é aprovado)
    R=Observação

  Fórmulas intocadas: N(menor preço), O(total menor), Q(total aut), totais linha 54
"""
import io, os
from datetime import date
import openpyxl

_TEMPLATE_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "template", "mapa_compras_template.xlsx"
)

DATA_START = 11
DATA_END   = 53
SUP_COLS   = ["I", "J", "K", "L"]
DATA_COLS  = ["E", "F", "G", "H", "I", "J", "K", "L", "P", "R"]


def generate_excel(
    items,
    supplier_names,
    numero_sequencial,
    filial,
    responsavel,
    data_compra,
    approved_supplier=None,
):
    """
    Gera o Excel do Mapa de Compras a partir do template.
    
    Args:
        items: Lista de itens normalizados
        supplier_names: Lista com nomes dos fornecedores
        numero_sequencial: Número sequencial do mapa
        filial: Nome da filial
        responsavel: Nome do responsável
        data_compra: Data da compra
        approved_supplier: Nome do fornecedor aprovado (opcional).
            Se fornecido, preenche a coluna P (Preço Autorizado) com os preços desse fornecedor.
    """
    wb = openpyxl.load_workbook(_TEMPLATE_PATH)
    ws = wb.active

    # Cabeçalho
    ws["C6"] = f" Número Sequencial: {numero_sequencial}"
    ws["F6"] = f"               Filial: {filial}"
    ws["H6"] = f"                                                                                                        Responsável pela compra: {responsavel}"
    ws["R6"] = f"Data: {data_compra.strftime('%d/%m/%Y')}"

    # Nomes dos fornecedores
    padded = (list(supplier_names) + ["", "", "", ""])[:4]
    for col, name in zip(SUP_COLS, padded):
        ws[f"{col}10"] = name.upper() if name else ""

    # Limpa linhas de dados
    for row in range(DATA_START, DATA_END + 1):
        for col in DATA_COLS:
            ws[f"{col}{row}"] = None

    # ── Trava de segurança: template suporta no máximo 43 itens ─────────────────
    MAX_ITEMS = DATA_END - DATA_START + 1   # = 43
    overflow  = len(items) > MAX_ITEMS
    items     = items[:MAX_ITEMS]

    def _safe_float(val, fallback=None):
        """Converte para float com segurança — retorna fallback se falhar."""
        if val is None:
            return fallback
        try:
            return float(val)
        except (ValueError, TypeError):
            return fallback

    # Preenche itens
    for i, item in enumerate(items):
        row = DATA_START + i
        ws[f"E{row}"] = (item.get("item") or "").upper()
        ws[f"F{row}"] = item.get("marca") or None
        ws[f"G{row}"] = _safe_float(item.get("quantidade"))
        ws[f"H{row}"] = item.get("unidade") or None

        fornecedores = item.get("fornecedores") or {}
        for col, fname in zip(SUP_COLS, padded):
            if not fname:
                ws[f"{col}{row}"] = None
                continue
            fdata = fornecedores.get(fname) or {}
            ws[f"{col}{row}"] = _safe_float(fdata.get("preco_unit"))

        # Preço Autorizado (coluna P) — se um fornecedor foi aprovado
        if approved_supplier and approved_supplier in fornecedores:
            fdata = fornecedores.get(approved_supplier) or {}
            ws[f"P{row}"] = _safe_float(fdata.get("preco_unit"))
        else:
            ws[f"P{row}"] = None

        # Observação consolidada
        obs_parts = []
        if item.get("observacao"):
            obs_parts.append(item["observacao"])
        for fname in padded:
            if not fname:
                continue
            fdata = fornecedores.get(fname) or {}
            if fdata.get("obs"):
                obs_parts.append(f"[{fname[:4]}] {fdata['obs']}")
        ws[f"R{row}"] = " | ".join(obs_parts) if obs_parts else None

    buf = io.BytesIO()
    wb.save(buf)
    # Retorna (bytes, aviso_de_overflow)
    warning = (
        f"⚠️ O mapa suporta no máximo 43 itens. "
        f"{len(items) + (len(items) if overflow else 0)} itens recebidos — "
        f"os últimos foram cortados."
        if overflow else None
    )
    return buf.getvalue(), warning
